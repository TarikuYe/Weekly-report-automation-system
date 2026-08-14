#!/usr/bin/env python3
"""
Weekly Report Consolidator
===========================
Merges every departmental .xlsx file in an input folder into a single,
polished master workbook — a "Details" sheet with all rows, and a
"Summary" sheet with per-department rollups.

Usage:
    python excel_consolidator.py [--config config.json]

Config file (config.json) fields:
    input_dir    -> folder containing department .xlsx files
    output_file  -> path for the consolidated master workbook
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference


def base_dir() -> Path:
    """Directory the app lives in — works whether run as a .py script or a
    PyInstaller-frozen .exe, and regardless of the OS's current working
    directory (which is unreliable when double-clicked from Explorer)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


# --------------------------------------------------------------------------- #
# Logging
# --------------------------------------------------------------------------- #
LOG_DIR = base_dir() / "logs"
LOG_DIR.mkdir(exist_ok=True)

_handlers: list[logging.Handler] = [
    logging.FileHandler(LOG_DIR / "app.log", encoding="utf-8"),
]
# sys.stdout is None when running as a .pyw / --windowed exe (no console).
# Adding a StreamHandler against None causes silent crashes inside logging.
if sys.stdout is not None:
    _handlers.append(logging.StreamHandler(sys.stdout))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=_handlers,
)
log = logging.getLogger("consolidator")


# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #
@dataclass
class Config:
    input_dir: Path
    output_file: Path

    @classmethod
    def load(cls, path: Path) -> "Config":
        if not path.exists():
            raise FileNotFoundError(f"Config file not found: {path}")
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        try:
            return cls(
                input_dir=Path(raw["input_dir"]),
                output_file=Path(raw["output_file"]),
            )
        except KeyError as exc:
            raise KeyError(f"Missing required config key: {exc}") from exc


# --------------------------------------------------------------------------- #
# Modern design tokens
# --------------------------------------------------------------------------- #
class Theme:
    FONT = "Aptos"  # falls back gracefully; swap to "Calibri" if unavailable
    FALLBACK_FONT = "Calibri"

    INK = "1E2430"          # near-black text
    MUTED = "6B7280"        # secondary text / borders
    ACCENT = "2D5BFF"       # primary brand blue
    ACCENT_DARK = "1B2A5C"  # header band
    SUCCESS = "1B8A5A"
    WARNING = "B7791F"
    DANGER = "C0392B"

    HEADER_FILL = PatternFill("solid", fgColor=ACCENT_DARK)
    BAND_FILL = PatternFill("solid", fgColor="F3F5FA")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

    THIN = Side(style="thin", color="E3E6EE")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    HEADER_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name=FONT, size=18, bold=True, color=ACCENT_DARK)
    SUBTITLE_FONT = Font(name=FONT, size=10, italic=True, color=MUTED)
    BODY_FONT = Font(name=FONT, size=10, color=INK)
    KPI_LABEL_FONT = Font(name=FONT, size=10, color=MUTED)
    KPI_VALUE_FONT = Font(name=FONT, size=20, bold=True, color=ACCENT_DARK)

    STATUS_FONT = {
        "complete": Font(name=FONT, size=10, bold=True, color=SUCCESS),
        "progress": Font(name=FONT, size=10, bold=True, color=WARNING),
        "blocked": Font(name=FONT, size=10, bold=True, color=DANGER),
    }

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")


# --------------------------------------------------------------------------- #
# Data loading
# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# Schema normalization — different departments name the same thing differently
# --------------------------------------------------------------------------- #
# Canonical column -> set of accepted header spellings (case/space-insensitive).
# Deliberately conservative: only merges columns we're confident mean the same
# thing. Anything not listed here is left untouched (visible, not lost).
COLUMN_ALIASES: dict[str, set[str]] = {
    "Date": {"date"},
    "Staff Name": {"staff name", "employee", "employee name", "engineer", "staff", "name"},
    "Department": {"department", "division", "dept", "team", "business unit"},
    "Project": {"project", "site / project", "site/project"},
    "Activity": {"activity", "inspection task", "task"},
    "Hours Worked": {"hours worked", "man hours", "hours", "time spent"},
    "Status": {"status"},
    "Remarks": {"remarks", "field notes", "notes", "comments"},
}

# A file only counts toward the department dashboard if, after normalization,
# it has *both* of these — otherwise its hours/department numbers can't be
# trusted alongside everyone else's.
REQUIRED_FOR_ROLLUP = {"Department", "Hours Worked"}


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Renames recognized column-name variants to a single canonical name."""
    rename_map = {}
    seen_canonical = set()
    for col in df.columns:
        key = str(col).strip().lower()
        for canonical, variants in COLUMN_ALIASES.items():
            if key in variants and canonical not in seen_canonical:
                rename_map[col] = canonical
                seen_canonical.add(canonical)
                break
    return df.rename(columns=rename_map)


# --------------------------------------------------------------------------- #
# Per-employee report format detection and parsing
# --------------------------------------------------------------------------- #
# Sentinel strings that mark the end of the daily-work data block.
_STOP_SENTINELS = {
    "key weekly outputs",
    "review and sign-off",
    "employee confirmation",
    "department head review",
}

# Labels used in the metadata block (rows 3-4 of each sheet).
# Maps the label text (lower-stripped) -> canonical field name we want.
_META_LABELS: dict[str, str] = {
    "employee name":   "Staff Name",
    "department":      "Department",
    "position":        "Position",
    "week start":      "Week Start",
    "week end":        "Week End",
    "report date":     "Report Date",
    "prepared by":     "Prepared By",
    "department head": "Department Head",
}


def _cell_str(cell) -> str:
    """Return a stripped string for a cell value, or '' if None/blank.

    openpyxl 3.1.5+ can return numpy arrays or plain lists for cells that
    contain array formulas (e.g. spill ranges).  pandas 3.x refuses to store
    those as strings.  We collapse any non-scalar to an empty string so the
    rest of the pipeline never sees them.
    """
    val = cell.value
    if val is None:
        return ""
    # Reject array-like values (numpy ndarray, list, tuple) — they come from
    # array-formula cells and are meaningless as text.
    if isinstance(val, (list, tuple)):
        return ""
    try:
        import numpy as np  # optional — only present when pandas is installed
        if isinstance(val, np.ndarray):
            return ""
    except ImportError:
        pass
    return str(val).strip()


def detect_report_format(wb) -> str:
    """Return 'employee_report' or 'flat_table'.

    An employee-report workbook has a recognisable metadata block near the top
    of its first sheet.  To avoid false-positives on generic dashboard reports
    that also contain words like "Department" or "Report Date", we require at
    least one hit from the *specific* labels that are unique to the weekly-hours
    employee template (employee name, week start, week end, department head),
    plus a total of at least two label hits overall.

    Handles two formats:
    - Label in one cell, value in the next cell
    - "Label: Value" combined in a single cell (new format)
    """
    # Labels that are specific to the weekly-hours employee-report template.
    # Generic labels like "department" or "report date" appear in many report
    # types and must not be used as sole discriminators.
    _SPECIFIC_LABELS = {"employee name", "week start", "week end", "department head"}

    ws = wb.worksheets[0]
    meta_hits = 0
    specific_hits = 0
    for row in ws.iter_rows(min_row=1, max_row=8):
        for cell in row:
            text = _cell_str(cell).lower()
            for label in _META_LABELS:
                if label in text:
                    meta_hits += 1
                    if label in _SPECIFIC_LABELS:
                        specific_hits += 1
                    break  # only count each cell once
    # Require at least one specific-label hit AND at least two total hits.
    if specific_hits >= 1 and meta_hits >= 2:
        return "employee_report"
    return "flat_table"


def _extract_metadata(ws) -> dict:
    """Scan the first 10 rows for label→value pairs and return a metadata dict.

    Handles two formats:
    1. Old format:  [Label] [Value] ... [Label] [Value]  (label and value in separate cells)
    2. New format:  "Label: Value" or "Label: Value | Label2: Value2"  (combined in one cell)
    """
    meta: dict[str, str] = {}
    
    for row in ws.iter_rows(min_row=1, max_row=10):
        cells = [c for c in row]
        for i, cell in enumerate(cells):
            text = _cell_str(cell)
            text_lower = text.lower()
            
            # --- Format 2: "Label: Value" in the same cell ---
            if ":" in text:
                # Split by pipe first (handles "Employee Name: X | Role: Y")
                segments = text.split("|")
                for segment in segments:
                    if ":" not in segment:
                        continue
                    label_part, _, value_part = segment.partition(":")
                    label_part = label_part.strip().lower()
                    value_part = value_part.strip()
                    
                    if label_part in _META_LABELS and value_part:
                        meta[_META_LABELS[label_part]] = value_part
            
            # --- Format 1: Label in this cell, value in next cell ---
            if text_lower in _META_LABELS:
                # value is the next non-empty cell on the same row
                for j in range(i + 1, len(cells)):
                    val = _cell_str(cells[j])
                    if val and ":" not in val:  # skip if it looks like another "Label: Value" pair
                        meta[_META_LABELS[text_lower]] = val
                        break
    
    return meta


def _find_data_header_row(ws) -> int | None:
    """Return the 1-based row number of the data header (contains 'Day' and 'Date').

    We scan up to row 20 to be safe — the template puts it around row 9.
    """
    for row in ws.iter_rows(min_row=1, max_row=20):
        texts = [_cell_str(c).lower() for c in row]
        if "day" in texts and "date" in texts:
            return row[0].row
    return None


def parse_employee_report_sheet(ws) -> pd.DataFrame | None:
    """Extract the daily work rows from a single formatted employee-report sheet.

    Returns a DataFrame with canonical columns, or None if the sheet has no
    recognisable data block (e.g. a cover or index sheet).

    Logic:
    - Extract metadata (Employee Name, Department, …) from the top block.
    - Find the data header row (contains "Day" and "Date").
    - Read every row below it until a stop sentinel is encountered.
    - Forward-fill Day and Date across Task 2 / Task 3 sub-rows.
    - Skip rows that have no meaningful content.
    - Attach metadata columns to every output row.
    """
    meta = _extract_metadata(ws)
    header_row_num = _find_data_header_row(ws)
    if header_row_num is None:
        return None

    # Build column-name list from the header row
    header_cells = list(ws.iter_rows(min_row=header_row_num,
                                      max_row=header_row_num))[0]
    col_names = [_cell_str(c) for c in header_cells]
    # Remove trailing empty column names
    while col_names and not col_names[-1]:
        col_names.pop()
    ncols = len(col_names)
    if ncols == 0:
        return None

    rows_out: list[dict] = []
    last_day: str = ""
    last_date = None

    for row in ws.iter_rows(min_row=header_row_num + 1):
        # Read only as many cells as we have column names
        vals = [_cell_str(c) for c in row[:ncols]]

        # Stop at section sentinels
        first_val = vals[0].lower() if vals else ""
        if any(s in first_val for s in _STOP_SENTINELS):
            break
        # Also stop if the whole row looks like a section header (one non-empty
        # value that contains a stop sentinel anywhere in the row)
        all_text = " ".join(v.lower() for v in vals if v)
        if any(s in all_text for s in _STOP_SENTINELS):
            break

        # Skip completely blank rows
        if not any(v for v in vals):
            continue

        # Build a dict for this row
        record: dict = {}
        for i, name in enumerate(col_names):
            record[name] = vals[i] if i < len(vals) else ""

        # Forward-fill Day and Date (Task 2 / Task 3 sub-rows leave them blank)
        day_val = record.get("Day", "").strip()
        date_val = record.get("Date", "").strip()

        if day_val and day_val.lower() not in ("task 2", "task 3"):
            last_day = day_val
            # Capture actual cell value for Date (may be a datetime object).
            # Guard against array-formula cells that return numpy arrays.
            date_col_idx = col_names.index("Date") if "Date" in col_names else -1
            if date_col_idx >= 0:
                raw_date = row[date_col_idx].value
                # Collapse array/list values — treat them as missing
                if isinstance(raw_date, (list, tuple)):
                    raw_date = None
                else:
                    try:
                        import numpy as np
                        if isinstance(raw_date, np.ndarray):
                            raw_date = None
                    except ImportError:
                        pass
                last_date = raw_date if raw_date is not None else date_val
        else:
            record["Day"] = last_day
            record["Date"] = last_date

        # Skip "Task 2" / "Task 3" rows that have no actual content beyond the label
        content_vals = [v for k, v in record.items()
                        if k not in ("Day", "Date") and str(v).strip()]
        if not content_vals:
            continue

        # Attach metadata
        for field, value in meta.items():
            record[field] = value

        rows_out.append(record)

    if not rows_out:
        return None

    df = pd.DataFrame(rows_out)

    # Normalise column names to canonical names
    df = normalize_columns(df)

    # Drop any columns with empty-string names — these are spacer columns in
    # the template (the first column is always blank in this format).
    df = df.loc[:, [c for c in df.columns if c.strip() != ""]]

    # Coerce Hours Worked to numeric (use loc to avoid chained-assignment warning)
    if "Hours Worked" in df.columns:
        df = df.copy()
        # Cast to object first so pandas 3.x StringDtype doesn't reject numeric values
        df["Hours Worked"] = df["Hours Worked"].astype(object)
        df.loc[:, "Hours Worked"] = pd.to_numeric(df["Hours Worked"], errors="coerce")
        
        # If Hours Worked is all NaN/empty, try to calculate it from entrance/leave times
        if df["Hours Worked"].isna().all() and "Office Entrance Hour" in df.columns and "Office Leave Hour" in df.columns:
            def calc_hours(row):
                try:
                    entrance = pd.to_datetime(str(row["Office Entrance Hour"]), format="%H:%M:%S", errors="coerce")
                    leave = pd.to_datetime(str(row["Office Leave Hour"]), format="%H:%M:%S", errors="coerce")
                    if pd.notna(entrance) and pd.notna(leave):
                        delta = leave - entrance
                        return delta.total_seconds() / 3600.0  # convert to hours
                except Exception:
                    pass
                return None
            
            calculated_hours = df.apply(calc_hours, axis=1)
            if calculated_hours.notna().any():
                df.loc[:, "Hours Worked"] = calculated_hours
                log.info(
                    "  Calculated Hours Worked from Office Entrance/Leave times (%d rows)",
                    calculated_hours.notna().sum()
                )

    # Coerce % Completion if present
    if "% Completion" in df.columns:
        pct = df["% Completion"].astype(str).str.replace("%", "", regex=False)
        df = df.copy()
        df["% Completion"] = df["% Completion"].astype(object)
        df.loc[:, "% Completion"] = pd.to_numeric(pct, errors="coerce")

    return df


def parse_employee_report_workbook(path: Path) -> tuple[pd.DataFrame, list[str]]:
    """Parse every sheet in a per-employee report workbook.

    Returns (combined_df, sheet_warnings).
    sheet_warnings lists sheets that were skipped (no data block found).
    """
    wb = load_workbook(path, data_only=True)
    frames: list[pd.DataFrame] = []
    warnings: list[str] = []

    for ws in wb.worksheets:
        df = parse_employee_report_sheet(ws)
        if df is None or df.empty:
            warnings.append(
                f"  Sheet '{ws.title}': no daily-work block found — skipped"
            )
            log.debug("Sheet '%s' in %s: no data block found, skipping", ws.title, path.name)
        else:
            log.debug("Sheet '%s' in %s: extracted %d rows", ws.title, path.name, len(df))
            frames.append(df)

    if not frames:
        raise ValueError(
            f"{path.name}: detected as employee-report format but no sheet "
            "yielded extractable daily-work rows."
        )

    combined = pd.concat(frames, ignore_index=True, sort=False)
    return combined, warnings


def detect_header_row(path: Path, max_scan: int = 15) -> tuple[int, pd.DataFrame | None]:
    """Finds the most header-like row in the first `max_scan` rows.

    Returns (header_row_index, preview_df_or_None). Returning the preview
    avoids reading the file twice — the caller can re-use it if header_row==0,
    or pass it back to pd.read_excel via skiprows for non-zero headers.

    Handles files that have title/blank rows above the real header (as
    opposed to well-formed reports where row 0 already is the header).
    Scores each row by how many non-null, short, string-like cells it has —
    exactly what a header row looks like, and what a title/blank row doesn't.
    """
    try:
        preview = pd.read_excel(path, header=None, nrows=max_scan)
    except Exception:
        return 0, None

    best_idx, best_score = 0, -1
    for i in range(len(preview)):
        row = preview.iloc[i]
        non_null = row.notna().sum()
        string_like = sum(
            isinstance(v, str) and 0 < len(v) < 40 for v in row if pd.notna(v)
        )
        score = non_null + string_like
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx, preview


def _resolve_dept_folder(input_dir: Path, folder_value: str) -> Path | None:
    """Return the absolute Path for a department's input folder.

    - Empty string  → None  (no folder configured).
    - Absolute path → used directly (external / shared network folder).
    - Relative name → joined to *input_dir* (legacy behaviour).
    """
    if not folder_value:
        return None
    p = Path(folder_value)
    if p.is_absolute():
        return p
    return input_dir / folder_value


def _collect_xlsx(folder: Path) -> list[Path]:
    """Return all non-temp .xlsx files inside *folder* recursively,
    skipping any ``archive`` sub-directories at any depth."""
    return sorted(
        p for p in folder.rglob("*.xlsx")
        if not p.name.startswith("~$")
        and "archive" not in [part.lower() for part in p.relative_to(folder).parts[:-1]]
    )


def discover_files(input_dir: Path, departments: list[dict] | None = None) -> list[Path]:
    """Find all .xlsx files to be merged.

    Two modes
    ---------
    departments provided (new mode):
        Each department's ``folder`` value is resolved by
        ``_resolve_dept_folder``.  If it is an absolute path the folder is
        scanned directly; otherwise it is treated as a sub-folder of
        *input_dir* (backward-compatible).  Root-level files directly under
        *input_dir* are always collected first so legacy flat layouts keep
        working.

    departments=None (legacy mode):
        Behaves exactly as before — scans *input_dir* root plus every
        non-archive direct sub-folder.

    The ``archive`` sub-folder is always excluded so previously-processed
    batches are never re-merged.
    """
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

    files: list[Path] = []

    # ── Root-level files (legacy flat layout, always included) ──────────
    files.extend(
        sorted(
            p for p in input_dir.glob("*.xlsx")
            if not p.name.startswith("~$")
        )
    )

    if departments is not None:
        # ── Per-department folders (may be external absolute paths) ──────
        seen_folders: set[Path] = set()
        for dept in departments:
            folder_value = dept.get("folder", "").strip()
            dept_folder  = _resolve_dept_folder(input_dir, folder_value)
            if dept_folder is None or not dept_folder.is_dir():
                continue
            if dept_folder in seen_folders:
                continue
            seen_folders.add(dept_folder)
            files.extend(_collect_xlsx(dept_folder))
    else:
        # ── Legacy mode: scan every non-archive direct sub-folder ────────
        for sub in sorted(input_dir.iterdir()):
            if not sub.is_dir() or sub.name.lower() == "archive":
                continue
            files.extend(_collect_xlsx(sub))

    if not files:
        raise FileNotFoundError(
            f"No .xlsx files found in {input_dir} or its department sub-folders"
        )
    return files


def load_and_merge(
    files: list[Path],
) -> tuple[pd.DataFrame, list[tuple[str, pd.DataFrame]], list[str], list[str]]:
    """Read every department file, skipping (not crashing on) bad ones.

    Returns
    -------
    matched_df
        Single merged DataFrame of every file/sheet that has both
        ``Department`` and ``Hours Worked`` (rollup-eligible rows).
        Carries ``_Rollup Eligible = True`` on every row.
    unmatched_sections
        List of ``(label, df)`` pairs — one entry per file (or per sheet for
        employee-report workbooks) — for sources whose layout did *not* match
        the expected schema.  These are written into the Details sheet as
        separate raw sections with a visual divider, so nothing is lost.
    failures
        File names that could not be read at all (I/O or parse errors).
    schema_warnings
        Human-readable notes about why each unmatched source was excluded.

    Two formats are handled automatically:
      flat_table       — a standard spreadsheet with one header row of data.
      employee_report  — a multi-sheet per-employee weekly template where each
                         sheet has a metadata block, a WEEKLY SUMMARY section,
                         and a daily-task data block (e.g. Mock_4departments.xlsx).
    """
    matched_frames: list[pd.DataFrame] = []
    unmatched_sections: list[tuple[str, pd.DataFrame]] = []
    failures: list[str] = []
    schema_warnings: list[str] = []

    for file in files:
        try:
            wb = load_workbook(file, data_only=True)
            fmt = detect_report_format(wb)

            if fmt == "employee_report":
                df, sheet_warns = parse_employee_report_workbook(file)
                df["Source File"] = file.name
                if sheet_warns:
                    schema_warnings.extend(sheet_warns)
                log.info(
                    "Loaded %-30s (%d rows from %d sheet(s), employee-report format)",
                    file.name, len(df), len(wb.worksheets),
                )
            else:
                # ---- flat_table path ----
                header_row, _ = detect_header_row(file)
                raw = pd.read_excel(file, header=None)
                raw.columns = range(len(raw.columns))
                header_vals = [
                    str(v).strip() if pd.notna(v) and not isinstance(v, (list, tuple)) else ""
                    for v in raw.iloc[header_row]
                ]
                df = raw.iloc[header_row + 1:].copy()
                df.columns = header_vals
                df = df.reset_index(drop=True)
                df.columns = [str(c).strip() for c in df.columns]
                # Coerce any array-typed cells to NaN so pandas 3.x doesn't reject them
                for col in df.columns:
                    mask = df[col].apply(lambda x: isinstance(x, (list, tuple)))
                    if mask.any():
                        df.loc[mask, col] = None
                df = normalize_columns(df)
                df["Source File"] = file.name
                log.info(
                    "Loaded %-30s (%d rows, %d cols, header row %d)",
                    file.name, len(df), len(df.columns), header_row,
                )

            # ---- decide: matched (rollup) or unmatched (raw section) ----
            missing = REQUIRED_FOR_ROLLUP - set(df.columns)
            has_data = (
                not missing
                and df["Department"].notna().any()
                and pd.to_numeric(df["Hours Worked"], errors="coerce").notna().any()
            )

            if has_data:
                df.loc[:, "_Rollup Eligible"] = True
                matched_frames.append(df)
            else:
                reason = (
                    f"missing column(s): {', '.join(sorted(missing))}"
                    if missing
                    else "Department / Hours Worked columns are empty"
                )
                schema_warnings.append(
                    f"{file.name}: {reason} — shown as raw section in Details"
                )
                log.warning(
                    "%s did not match the weekly-hours schema (%s) — "
                    "will appear as raw section in Details",
                    file.name, reason,
                )
                # One section per source file (label = filename)
                unmatched_sections.append((file.name, df))

        except Exception as exc:  # noqa: BLE001
            failures.append(f"{file.name}: {exc}")
            log.warning("Skipped %s due to read error: %s", file.name, exc)

    if not matched_frames and not unmatched_sections:
        raise RuntimeError("Every input file failed to load; nothing to consolidate.")

    if matched_frames:
        matched_df = pd.concat(matched_frames, ignore_index=True, sort=False)
    else:
        # No matched files — build an empty placeholder so downstream code
        # doesn't need to special-case None.
        matched_df = pd.DataFrame()

    return matched_df, unmatched_sections, failures, schema_warnings


# --------------------------------------------------------------------------- #
# Styling helpers
# --------------------------------------------------------------------------- #
def style_header_row(ws, ncols: int, row: int = 1) -> None:
    ws.row_dimensions[row].height = 26
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Theme.HEADER_FONT
        cell.fill = Theme.HEADER_FILL
        cell.alignment = Theme.CENTER
        cell.border = Theme.BORDER


def style_data_rows(ws, first_row: int, last_row: int, ncols: int,
                     status_col: Optional[int] = None,
                     center_cols: Optional[set[int]] = None) -> None:
    center_cols = center_cols or set()
    for row in range(first_row, last_row + 1):
        ws.row_dimensions[row].height = 20
        banded = (row - first_row) % 2 == 1
        fill = Theme.BAND_FILL if banded else Theme.WHITE_FILL

        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Theme.BODY_FONT
            cell.fill = fill
            cell.border = Theme.BORDER
            cell.alignment = Theme.CENTER if col in center_cols else Theme.LEFT

            if status_col and col == status_col and cell.value:
                val = str(cell.value).strip().lower()
                for key, font in Theme.STATUS_FONT.items():
                    if key in val:
                        cell.font = font
                        break


def autofit_columns(ws, min_width: int = 12, max_width: int = 48) -> None:
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_letter = get_column_letter(cell.column)
            widths[col_letter] = max(widths.get(col_letter, 0), len(str(cell.value)))
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(width + 4, max_width))


def add_table(ws, ncols: int, nrows: int, name: str) -> None:
    """Registers a native Excel table so filters/sorting work out of the box."""
    last_col = get_column_letter(ncols)
    ref = f"A1:{last_col}{nrows}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=False,  # we hand-roll our own banding
    )
    ws.add_table(table)


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #
def write_table_block(ws, start_row: int, table_df: pd.DataFrame,
                       center_cols: Optional[set[int]] = None) -> int:
    """Writes a small styled table starting at start_row. Returns the row after it."""
    headers = list(table_df.columns)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    for r, record in enumerate(table_df.itertuples(index=False), start=start_row + 1):
        for c, val in enumerate(record, start=1):
            ws.cell(row=r, column=c, value=val)

    ncols = len(headers)
    nrows = len(table_df) + 1
    style_header_row(ws, ncols, row=start_row)
    style_data_rows(
        ws, start_row + 1, start_row + nrows - 1, ncols,
        center_cols=center_cols or set(range(2, ncols + 1)),
    )
    return start_row + nrows


def add_bar_chart(ws, anchor: str, title: str, y_title: str, x_title: str,
                   value_col: int, cat_col: int, header_row: int, last_row: int) -> None:
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    data = Reference(ws, min_col=value_col, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 16, 8
    ws.add_chart(chart, anchor)


# --------------------------------------------------------------------------- #
# Details-sheet section divider helpers
# --------------------------------------------------------------------------- #
# Visual style for the divider banner row that introduces each raw section.
_DIVIDER_FILL   = PatternFill("solid", fgColor="1B2A5C")   # same as ACCENT_DARK
_DIVIDER_FONT   = Font(name=Theme.FONT, size=11, bold=True, color="FFFFFF")
_DIVIDER_HEIGHT = 24

# Style for the sub-header row that labels each raw section's columns.
_RAW_HDR_FILL   = PatternFill("solid", fgColor="3A5490")
_RAW_HDR_FONT   = Font(name=Theme.FONT, size=10, bold=True, color="FFFFFF")
_RAW_HDR_HEIGHT = 20

# Style for raw data rows inside an unmatched section.
_RAW_BAND_FILL  = PatternFill("solid", fgColor="EEF1FA")
_RAW_BODY_FONT  = Font(name=Theme.FONT, size=9, color=Theme.INK)
_RAW_BODY_HEIGHT = 18


def _write_divider_row(ws, row: int, label: str, ncols: int) -> None:
    """Write a full-width banner row labelled with *label* at *row*.

    The first cell gets the label text; all cells in the row are merged,
    filled with the dark accent colour, and given white bold text so it
    reads as a clear visual separator.
    """
    ws.row_dimensions[row].height = _DIVIDER_HEIGHT

    # Write the label in column A, leave the rest empty before merging.
    cell = ws.cell(row=row, column=1, value=f"  📄  {label}")
    cell.font   = _DIVIDER_FONT
    cell.fill   = _DIVIDER_FILL
    cell.alignment = Theme.LEFT

    # Fill every other cell in the band (merge alone doesn't colour them).
    for col in range(2, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _DIVIDER_FILL

    if ncols > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=ncols,
        )


def _write_raw_section(ws, start_row: int, label: str,
                        section_df: pd.DataFrame,
                        master_ncols: int) -> int:
    """Write one unmatched source into the worksheet starting at *start_row*.

    Layout
    ------
    start_row     : divider banner  (full-width, dark blue, white bold text)
    start_row + 1 : column-header row for this section's own columns
    start_row + 2 … : data rows, lightly banded
    (blank gap row appended at the end)

    Returns the next available row number after the gap.

    *master_ncols* is the width of the matched table above (used to size
    the banner so it spans the same width even when this section is narrower).
    """
    cols = list(section_df.columns)
    ncols = max(len(cols), master_ncols)   # banner spans at least as wide as the matched table

    # --- divider banner ---
    _write_divider_row(ws, start_row, label, ncols)
    row = start_row + 1

    # --- column-header row ---
    ws.row_dimensions[row].height = _RAW_HDR_HEIGHT
    for ci, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=ci, value=col_name)
        cell.font      = _RAW_HDR_FONT
        cell.fill      = _RAW_HDR_FILL
        cell.alignment = Theme.CENTER
        cell.border    = Theme.BORDER
    row += 1

    # --- data rows ---
    for ri, record in enumerate(section_df.itertuples(index=False, name=None)):
        ws.row_dimensions[row].height = _RAW_BODY_HEIGHT
        fill = _RAW_BAND_FILL if ri % 2 == 1 else Theme.WHITE_FILL
        for ci, val in enumerate(record, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font      = _RAW_BODY_FONT
            cell.fill      = fill
            cell.alignment = Theme.LEFT
            cell.border    = Theme.BORDER
        row += 1

    # --- blank gap row before the next section ---
    ws.row_dimensions[row].height = 10
    row += 1

    return row


def build_details_sheet(
    wb,
    matched_df: pd.DataFrame,
    unmatched_sections: list[tuple[str, pd.DataFrame]],
) -> None:
    """Build the Details sheet.

    Structure
    ---------
    1. If there are any matched (rollup-eligible) rows:
       • One unified styled table with an Excel AutoFilter, freeze pane,
         and the "In Dashboard?" flag column.
    2. For every unmatched source:
       • A dark banner row with the source file name.
       • A sub-header row showing that source's own column names.
       • Its raw data rows, lightly banded.
       • A small gap before the next section.
    """
    ws = wb.active
    ws.title = "Details"
    ws.sheet_view.showGridLines = False

    matched_ncols = 0   # track width of matched table for banner sizing

    # ------------------------------------------------------------------ #
    # Part 1 — matched / rollup-eligible rows
    # ------------------------------------------------------------------ #
    if not matched_df.empty:
        display_df = matched_df.copy()
        display_df = display_df.rename(columns={"_Rollup Eligible": "In Dashboard?"})
        if "In Dashboard?" in display_df.columns:
            in_dash = (
                display_df.pop("In Dashboard?")
                .map({True: "Yes", False: "No"})
                .astype(str)
            )
            display_df.insert(len(display_df.columns), "In Dashboard?", in_dash)

        matched_ncols = len(display_df.columns)

        for row in dataframe_rows(display_df):
            ws.append(row)

        nrows = len(display_df) + 1   # includes header
        style_header_row(ws, matched_ncols)

        status_col   = None
        center_cols  = set()
        for idx, col_name in enumerate(display_df.columns, start=1):
            if col_name == "Status":
                status_col = idx
                center_cols.add(idx)
            elif col_name in ("Hours Worked", "Department", "In Dashboard?"):
                center_cols.add(idx)

        style_data_rows(
            ws, 2, nrows, matched_ncols,
            status_col=status_col, center_cols=center_cols,
        )
        ws.freeze_panes = "A2"
        add_table(ws, matched_ncols, nrows, "DetailsTable")

        next_row = nrows + 2   # one blank gap row between table and first section
    else:
        next_row = 1   # sheet is empty — start dividers from row 1

    # ------------------------------------------------------------------ #
    # Part 2 — unmatched sources, one section each
    # ------------------------------------------------------------------ #
    for label, section_df in unmatched_sections:
        # Drop internal bookkeeping columns the user doesn't need to see.
        display_sec = section_df.drop(
            columns=[c for c in ("_Rollup Eligible", "Source File") if c in section_df.columns],
            errors="ignore",
        )
        next_row = _write_raw_section(
            ws, next_row, label, display_sec, master_ncols=matched_ncols,
        )

    autofit_columns(ws)
    ws.page_setup.orientation    = "landscape"
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_summary_sheet(wb, matched_df: pd.DataFrame, schema_warnings: list[str], failures: list[str]) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Summary only reflects rollup-eligible (matched) data.
    df = matched_df  # local alias for brevity
    rollup_df = df[df["_Rollup Eligible"]] if "_Rollup Eligible" in df.columns and not df.empty else df

    # --- Title ---
    ws["A1"] = "Weekly Report — Executive Dashboard"
    ws["A1"].font = Theme.TITLE_FONT
    ws.merge_cells("A1:H1")

    dept_count = rollup_df["Department"].nunique() if not rollup_df.empty else 0
    source_count = df["Source File"].nunique() if not df.empty and "Source File" in df.columns else 0
    ws["A2"] = (
        f"Consolidated from {source_count} matched source file(s) · "
        f"{len(df)} total rows · {dept_count} department(s) in dashboard"
    )
    ws["A2"].font = Theme.SUBTITLE_FONT
    ws.merge_cells("A2:H2")

    # --- KPI cards ---
    total_hours = pd.to_numeric(rollup_df["Hours Worked"], errors="coerce").sum() if not rollup_df.empty and "Hours Worked" in rollup_df.columns else 0
    staff_count = rollup_df["Staff Name"].nunique() if "Staff Name" in rollup_df.columns and not rollup_df.empty else None
    completed = None
    if "Status" in rollup_df.columns and not rollup_df.empty:
        completed = (rollup_df["Status"].astype(str).str.lower().str.contains("complet")).mean() * 100

    kpis = [
        ("Total Rows", len(df)),
        ("Total Hours Logged", f"{total_hours:,.1f}"),
        ("Departments Reporting", dept_count),
    ]
    if staff_count is not None:
        kpis.append(("Staff Reporting", staff_count))
    if staff_count:
        kpis.append(("Avg Hours / Person", f"{total_hours / staff_count:,.1f}"))
    if completed is not None:
        kpis.append(("Completion Rate", f"{completed:,.0f}%"))

    col = 1
    for label, value in kpis:
        ws.cell(row=4, column=col, value=label).font = Theme.KPI_LABEL_FONT
        ws.cell(row=5, column=col, value=value).font = Theme.KPI_VALUE_FONT
        col += 2

    row_cursor = 8

    # --- Department comparison table + chart ---
    if not rollup_df.empty and "Department" in rollup_df.columns:
        agg_kwargs = {"Rows": ("Department", "count")}
        if "Hours Worked" in rollup_df.columns:
            agg_kwargs["Total Hours"] = ("Hours Worked", "sum")
        if "Staff Name" in rollup_df.columns:
            agg_kwargs["Staff"] = ("Staff Name", "nunique")

        dept_table = rollup_df.groupby("Department").agg(**agg_kwargs).reset_index()

        if "Status" in rollup_df.columns:
            completion = (
                rollup_df.assign(_done=rollup_df["Status"].astype(str).str.lower().str.contains("complet"))
                .groupby("Department")["_done"].mean().mul(100).round(0)
            )
            dept_table = dept_table.assign(**{"Completed %": dept_table["Department"].map(completion)})

        if "Total Hours" in dept_table.columns:
            dept_table = dept_table.assign(**{"Avg Hrs/Row": (dept_table["Total Hours"] / dept_table["Rows"]).round(1)})
            dept_table = dept_table.sort_values("Total Hours", ascending=False)
        else:
            dept_table = dept_table.sort_values("Rows", ascending=False)

        ws.cell(row=row_cursor, column=1, value="Department Comparison").font = Theme.SUBTITLE_FONT
        row_cursor += 1
        table_start = row_cursor
        row_cursor = write_table_block(ws, table_start, dept_table)

        if "Total Hours" in dept_table.columns:
            value_col = list(dept_table.columns).index("Total Hours") + 1
            add_bar_chart(
                ws, f"A{row_cursor + 1}", "Total Hours by Department", "Hours", "Department",
                value_col=value_col, cat_col=1, header_row=table_start, last_row=row_cursor - 1,
            )
            row_cursor += 18  # leave room for the chart before the next block

    # --- Status breakdown ---
    if not rollup_df.empty and "Status" in rollup_df.columns:
        status_table = (
            rollup_df["Status"].astype(str).str.strip().value_counts()
            .rename_axis("Status").reset_index(name="Count")
        )
        status_table = status_table.assign(**{"% of Total": (status_table["Count"] / status_table["Count"].sum() * 100).round(0)})

        ws.cell(row=row_cursor, column=1, value="Status Breakdown (All Departments)").font = Theme.SUBTITLE_FONT
        row_cursor += 1
        table_start = row_cursor
        row_cursor = write_table_block(ws, table_start, status_table)
        add_bar_chart(
            ws, f"A{row_cursor + 1}", "Work Status Breakdown", "Count", "Status",
            value_col=2, cat_col=1, header_row=table_start, last_row=row_cursor - 1,
        )
        row_cursor += 18

    # --- Hours by project ---
    if not rollup_df.empty and "Project" in rollup_df.columns and "Hours Worked" in rollup_df.columns \
            and rollup_df["Project"].notna().any():
        project_table = (
            rollup_df.groupby("Project")["Hours Worked"].sum()
            .sort_values(ascending=False).round(1)
            .rename_axis("Project").reset_index(name="Total Hours")
        )
        ws.cell(row=row_cursor, column=1, value="Hours by Project (Cross-Department)").font = Theme.SUBTITLE_FONT
        row_cursor += 1
        table_start = row_cursor
        row_cursor = write_table_block(ws, table_start, project_table)
        add_bar_chart(
            ws, f"A{row_cursor + 1}", "Hours by Project", "Hours", "Project",
            value_col=2, cat_col=1, header_row=table_start, last_row=row_cursor - 1,
        )
        row_cursor += 18

    # --- Data quality callout ---
    if schema_warnings or failures:
        ws.cell(row=row_cursor, column=1, value="⚠ Data Quality Notes").font = Font(
            name=Theme.FONT, size=11, bold=True, color=Theme.WARNING,
        )
        row_cursor += 1
        note_fill = PatternFill("solid", fgColor="FFF6E5")
        for msg in failures + schema_warnings:
            cell = ws.cell(row=row_cursor, column=1, value=f"• {msg}")
            cell.font = Font(name=Theme.FONT, size=9, color=Theme.INK)
            cell.fill = note_fill
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
            row_cursor += 1
        row_cursor += 1
        note = ws.cell(
            row=row_cursor, column=1,
            value="Files marked above are shown as raw sections in the Details tab.",
        )
        note.font = Theme.SUBTITLE_FONT
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)

    autofit_columns(ws)


def dataframe_rows(df: pd.DataFrame):
    """Yields header row then data rows for openpyxl.append()."""
    yield list(df.columns)
    for record in df.itertuples(index=False, name=None):
        yield list(record)


# --------------------------------------------------------------------------- #
# Archive helpers
# --------------------------------------------------------------------------- #
def archive_inputs(input_dir: Path, departments: list[dict] | None = None) -> Path:
    """Move every .xlsx file in *input_dir* (and any configured department
    folders) into timestamped archive sub-folders.

    **NEW BEHAVIOR**: Each department's files are archived inside that
    department's own folder at <dept_folder>/archive/YYYY-MM-DD_HH-MM-SS/.
    Root-level files are archived under <input_dir>/archive/YYYY-MM-DD_HH-MM-SS/.

    This allows each department head to see their own submission history
    independently, and the Deputy General Manager can see the root-level
    archive for global files.

    Archive layout:
        <input_dir>/archive/YYYY-MM-DD_HH-MM-SS/
            file_at_root.xlsx   ← root-level files only

        <input_dir>/Contract/archive/YYYY-MM-DD_HH-MM-SS/
            contract_rpt.xlsx   ← Contract dept files

        <input_dir>/Design/archive/YYYY-MM-DD_HH-MM-SS/
            design_rpt.xlsx     ← Design dept files

        ...

    Returns the Path of the root-level archive sub-folder (for backward
    compatibility with GUI/logging).
    Raises RuntimeError if no files are found to archive.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    # Track all archive operations for logging
    archived_count = 0
    archive_locations: list[Path] = []

    # ── Archive root-level files ─────────────────────────────────────────
    root_files: list[Path] = []
    for p in sorted(input_dir.glob("*.xlsx")):
        if not p.name.startswith("~$"):
            root_files.append(p)

    if root_files:
        root_archive_dir = input_dir / "archive" / timestamp
        root_archive_dir.mkdir(parents=True, exist_ok=True)
        for src in root_files:
            dest = root_archive_dir / src.name
            shutil.move(str(src), str(dest))
            log.info("Archived: %s  →  %s", src, dest)
            archived_count += 1
        archive_locations.append(root_archive_dir)
        log.info("Archived %d root-level file(s) to %s", len(root_files), root_archive_dir)

    # ── Archive department files inside their own folders ────────────────
    if departments is not None:
        seen_folders: set[Path] = set()
        for dept in departments:
            folder_value = dept.get("folder", "").strip()
            dept_folder  = _resolve_dept_folder(input_dir, folder_value)
            if dept_folder is None or not dept_folder.is_dir():
                continue
            if dept_folder in seen_folders:
                continue
            seen_folders.add(dept_folder)

            dept_name = dept.get("name", dept_folder.name)

            # Collect all .xlsx files in this department folder (recursive)
            dept_files: list[Path] = []
            for p in sorted(dept_folder.rglob("*.xlsx")):
                if not p.name.startswith("~$"):
                    # Skip files already inside an archive subdirectory
                    try:
                        rel_parts = p.relative_to(dept_folder).parts[:-1]
                    except ValueError:
                        rel_parts = ()
                    if "archive" in [part.lower() for part in rel_parts]:
                        continue
                    dept_files.append(p)

            if dept_files:
                # Archive inside this department's own archive subfolder
                dept_archive_dir = dept_folder / "archive" / timestamp
                dept_archive_dir.mkdir(parents=True, exist_ok=True)
                for src in dept_files:
                    # Preserve relative path within the department folder
                    rel_path = src.relative_to(dept_folder)
                    dest = dept_archive_dir / rel_path
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(src), str(dest))
                    log.info("Archived: %s  →  %s", src, dest)
                    archived_count += 1
                archive_locations.append(dept_archive_dir)
                log.info(
                    "Archived %d file(s) from dept '%s' to %s",
                    len(dept_files), dept_name, dept_archive_dir,
                )
    else:
        # Legacy mode — scan direct sub-folders of input_dir
        for sub in sorted(input_dir.iterdir()):
            if not sub.is_dir() or sub.name.lower() == "archive":
                continue
            
            sub_files: list[Path] = []
            for p in sorted(sub.rglob("*.xlsx")):
                if not p.name.startswith("~$"):
                    # Skip files already inside an archive subdirectory
                    try:
                        rel_parts = p.relative_to(sub).parts[:-1]
                    except ValueError:
                        rel_parts = ()
                    if "archive" in [part.lower() for part in rel_parts]:
                        continue
                    sub_files.append(p)

            if sub_files:
                # Archive inside this subfolder's own archive directory
                sub_archive_dir = sub / "archive" / timestamp
                sub_archive_dir.mkdir(parents=True, exist_ok=True)
                for src in sub_files:
                    rel_path = src.relative_to(sub)
                    dest = sub_archive_dir / rel_path
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(src), str(dest))
                    log.info("Archived: %s  →  %s", src, dest)
                    archived_count += 1
                archive_locations.append(sub_archive_dir)
                log.info(
                    "Archived %d file(s) from subfolder '%s' to %s",
                    len(sub_files), sub.name, sub_archive_dir,
                )

    if archived_count == 0:
        raise RuntimeError(f"No .xlsx files to archive in {input_dir}")

    log.info(
        "Archived %d input file(s) total across %d location(s)",
        archived_count, len(archive_locations),
    )
    
    # Return the root-level archive dir for backward compatibility (GUI looks for this)
    # If no root-level files were archived, return the first dept archive location
    root_archive_dir = input_dir / "archive" / timestamp
    if not root_archive_dir.exists() and archive_locations:
        return archive_locations[0]
    return root_archive_dir


# --------------------------------------------------------------------------- #
# Pipeline
# --------------------------------------------------------------------------- #
def _execute(config: Config, departments: list[dict] | None = None) -> None:
    """Core pipeline logic, shared by both entry points below."""
    log.info("Starting Weekly Report Consolidation pipeline...")

    files = discover_files(config.input_dir, departments)
    log.info("Found %d department file(s) in %s", len(files), config.input_dir)

    matched_df, unmatched_sections, failures, schema_warnings = load_and_merge(files)

    matched_rows = len(matched_df) if not matched_df.empty else 0
    log.info(
        "Merged %d matched row(s) from %d file(s); %d unmatched section(s)",
        matched_rows, len(files) - len(failures), len(unmatched_sections),
    )
    if failures:
        log.warning("%d file(s) failed to load:\n  - %s", len(failures), "\n  - ".join(failures))
    if schema_warnings:
        log.warning(
            "%d file(s) shown as raw sections (schema mismatch):\n  - %s",
            len(schema_warnings), "\n  - ".join(schema_warnings),
        )

    if not matched_df.empty and "Hours Worked" in matched_df.columns:
        rollup_mask = matched_df.get("_Rollup Eligible", pd.Series(True, index=matched_df.index))
        total_hours = pd.to_numeric(
            matched_df.loc[rollup_mask, "Hours Worked"], errors="coerce"
        ).sum()
        log.info("Total company hours logged (dashboard-eligible rows): %.1f", total_hours)

    config.output_file.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook
    wb = Workbook()
    build_details_sheet(wb, matched_df, unmatched_sections)
    build_summary_sheet(wb, matched_df, schema_warnings, failures)
    wb.move_sheet("Summary", offset=-1)  # Summary first, Details second
    wb.active = wb.sheetnames.index("Summary")

    # Write to a temp file first, then atomically replace the target.
    # This avoids a PermissionError when the previous output is still open
    # in Excel (which locks the file on Windows).
    import tempfile
    tmp_path: Path | None = None
    try:
        fd, tmp_str = tempfile.mkstemp(
            suffix=".xlsx", dir=config.output_file.parent, prefix=".~tmp_"
        )
        os.close(fd)
        tmp_path = Path(tmp_str)
        wb.save(tmp_path)
        # Replace the target — on Windows shutil.move handles the overwrite
        shutil.move(str(tmp_path), str(config.output_file))
        tmp_path = None   # successfully moved; no cleanup needed
    except PermissionError:
        # The target file is open in another application.
        # Fall back to a timestamped alternative name so we don't lose the data.
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        alt_path = config.output_file.with_stem(
            f"{config.output_file.stem}_{ts}"
        )
        if tmp_path and tmp_path.exists():
            shutil.move(str(tmp_path), str(alt_path))
            tmp_path = None
        else:
            wb.save(alt_path)
        log.warning(
            "Output file is open in another application — saved to %s instead.\n"
            "  Close the file in Excel and rename it if needed.",
            alt_path,
        )
        config.output_file = alt_path   # update for the archive/log step below
    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass

    log.info("Consolidated master report saved to %s", config.output_file)

    # ── Per-department master Excel files (dept-head access only) ────────
    # Generates one password-protected workbook per department and stores the
    # results so email_sender can attach them.  Failures are logged but never
    # abort the pipeline — the global report is already safely written.
    dept_master_results: list = []
    if departments:
        try:
            import dept_master as _dm  # noqa: PLC0415
            from email_sender import week_label_for  # noqa: PLC0415
            wl = week_label_for()
            dept_master_results = _dm.generate_dept_masters(
                matched_df=matched_df,
                unmatched_sections=unmatched_sections,
                departments=departments,
                output_dir=config.output_file.parent,
                week_label=wl,
                input_dir=config.input_dir,
            )
            ok  = [r for r in dept_master_results if not r.error]
            err = [r for r in dept_master_results if r.error]
            if ok:
                log.info(
                    "Dept master files generated (%d): %s",
                    len(ok),
                    ", ".join(r.dept_name for r in ok),
                )
            if err:
                log.warning(
                    "Dept master generation failed for: %s",
                    ", ".join(f"{r.dept_name} ({r.error})" for r in err),
                )

            # ── Warn when a dept master was written but contains no data ─
            # This catches the "archive races ahead of dept_master" class of
            # bugs: if source files were submitted for a department but the
            # resulting workbook ended up empty, something went wrong in the
            # matching step and the dept head would receive a blank report.
            if not matched_df.empty and "Source File" in matched_df.columns and "Department" in matched_df.columns:
                import dept_master as _dm_check  # noqa: PLC0415
                for result in ok:
                    # Find the dept config entry for this result
                    dept_cfg = next(
                        (d for d in departments if d.get("name", "") == result.dept_name),
                        None,
                    )
                    if dept_cfg is None:
                        continue
                    # Check whether any rows in matched_df belong to this dept
                    source_files = _dm_check._source_files_for_dept_from_data(
                        dept_cfg, matched_df, unmatched_sections, config.input_dir
                    )
                    dept_col_mask = matched_df["Department"].apply(
                        lambda v, dn=result.dept_name: _dm_check._dept_matches(dn, str(v))
                        if pd.notna(v) else False
                    )
                    folder_mask = (
                        matched_df["Source File"].apply(
                            lambda f, sf=source_files: Path(str(f)).name in sf
                            if pd.notna(f) else False
                        )
                        if source_files else pd.Series(False, index=matched_df.index)
                    )
                    has_data = (dept_col_mask | folder_mask).any()
                    if not has_data:
                        log.warning(
                            "Dept master for '%s' was generated but contains NO DATA — "
                            "no rows in the merged dataset matched this department.  "
                            "Check that the Department column in submitted files matches "
                            "the config name '%s' (current values in data: %s).",
                            result.dept_name,
                            result.dept_name,
                            ", ".join(
                                f"'{v}'" for v in matched_df["Department"]
                                .dropna().astype(str).unique()
                            ),
                        )
        except Exception as exc:  # noqa: BLE001
            log.warning("Dept master generation skipped: %s", exc)

    # ── Email dept masters to department heads ───────────────────────────
    if dept_master_results and departments:
        try:
            # Load config for SMTP settings (needed by email_sender).
            # We attempt to read config.json from the same directory as the
            # output file; fall back gracefully if it doesn't exist.
            _cfg_path = base_dir() / "config.json"
            if _cfg_path.exists():
                with open(_cfg_path, "r", encoding="utf-8") as _f:
                    _cfg = json.load(_f)
                if _cfg.get("email", {}).get("enabled", False):
                    from email_sender import send_dept_master  # noqa: PLC0415
                    for result in dept_master_results:
                        if result.error or not result.dept_head_email:
                            continue
                        ok_send, msg_send = send_dept_master(_cfg, result)
                        if ok_send:
                            log.info(
                                "Dept master emailed to %s (%s)",
                                result.dept_head_email, result.dept_name,
                            )
                        else:
                            log.warning(
                                "Failed to email dept master to %s: %s",
                                result.dept_head_email, msg_send,
                            )
                else:
                    log.info(
                        "Email automation disabled — dept masters saved locally only. "
                        "Path: %s",
                        config.output_file.parent / "dept_masters",
                    )
        except Exception as exc:  # noqa: BLE001
            log.warning("Dept master email step skipped: %s", exc)

    # Archive input files so the next week starts with a clean input folder.
    try:
        archive_dir = archive_inputs(config.input_dir, departments)
        log.info("Input folder is now clean and ready for next week's reports.")
    except Exception as exc:  # noqa: BLE001
        log.warning("Archiving skipped: %s", exc)

    log.info("Pipeline complete. 🎉")


def run_pipeline(config_path: Path) -> None:
    """Entry point for CLI use: reads input_dir/output_file from a JSON config file."""
    config = Config.load(config_path)
    # Load departments for per-dept external folder support
    with open(config_path, "r", encoding="utf-8") as _f:
        _raw = json.load(_f)
    departments = _raw.get("departments", None)
    _execute(config, departments)


def run_pipeline_from_paths(input_dir: str, output_file: str, departments: list[dict] | None = None) -> None:
    """Entry point for programmatic/GUI use: pass paths directly, no config file needed.

    Pass *departments* (the list from config.json) to enable per-department
    external folder scanning.  When omitted the legacy sub-folder scan is used.
    """
    config = Config(input_dir=Path(input_dir), output_file=Path(output_file))
    _execute(config, departments)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Consolidate department Excel reports.")
    parser.add_argument(
        "--config", type=Path, default=Path("config.json"),
        help="Path to the JSON config file (default: config.json)",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        run_pipeline(args.config)
    except Exception as exc:  # noqa: BLE001 - top-level guard, fully logged
        log.error("Pipeline failed: %s", exc, exc_info=True)
        sys.exit(1)