"""
dept_master.py
==============
Generates a **per-department** master Excel workbook from the already-merged
DataFrame produced by the main pipeline, then optionally password-protects it
so only the department head can open it.

Public API
----------
generate_dept_masters(matched_df, unmatched_sections, departments, output_dir,
                      week_label)
    Creates one .xlsx per department that has a dept_head_email configured.
    Returns a list of DeptMasterResult named-tuples.

password_protect(src_path, dest_path, password)
    Wraps the file in an OOXML password container using msoffcrypto-tool.
    Falls back to a plain copy if the library is unavailable.

Matching strategy
-----------------
The ``Department`` column inside each employee's Excel file is free-text and
often differs from the config ``name``.  Common mismatches:

    Config name              File value
    ─────────────────────    ────────────────────────────────
    "Design"                 "Design"               ← exact
    "Office_Engineering"     "Office Engineering"   ← underscores vs spaces
    "Supervision"            "Supervision Dept"     ← extra suffix
    "Contract"               "Procurement & Contract Administration"  ← very different

We apply a three-tier fallback:

  1. Exact match (case-insensitive, stripped).
  2. Token-overlap match: split both sides on non-alphanumeric chars; if every
     token in the shorter name appears in the longer name → match.
  3. Source-file origin: if a file came from this department's configured input
     folder (stored in ``Source File`` column), those rows belong here regardless
     of what the Department column says.

For unmatched sections (files that didn't have a Department column) the same
folder-origin check is used so raw sections are still included in the right
dept workbook.
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import NamedTuple

import pandas as pd

log = logging.getLogger("dept_master")


# ---------------------------------------------------------------------------
# Result type returned to callers
# ---------------------------------------------------------------------------
class DeptMasterResult(NamedTuple):
    dept_name: str
    output_path: Path          # final file (encrypted if password was set)
    encrypted: bool
    dept_head_email: str
    error: str                 # non-empty when this dept failed


# ---------------------------------------------------------------------------
# Password protection
# ---------------------------------------------------------------------------
def password_protect(src_path: Path, dest_path: Path, password: str) -> bool:
    """Encrypt *src_path* with *password* and write to *dest_path*.

    Uses msoffcrypto-tool (pip install msoffcrypto-tool).
    Returns True on success, False if the library is missing.
    """
    try:
        import msoffcrypto  # type: ignore
    except ImportError:
        log.warning(
            "msoffcrypto-tool not installed — %s will NOT be password-protected.  "
            "Run: pip install msoffcrypto-tool",
            src_path.name,
        )
        return False

    try:
        with open(src_path, "rb") as f_in:
            office_file = msoffcrypto.OfficeFile(f_in)
            office_file.encrypt(password, dest_path.open("wb"))
        log.info("Encrypted %s → %s", src_path.name, dest_path.name)
        return True
    except Exception as exc:
        log.error("Encryption failed for %s: %s", src_path.name, exc)
        return False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_name(name: str) -> str:
    """Convert a department name to a filesystem-safe stem."""
    safe = "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in name)
    return safe.strip("_") or "dept"


def _tokenize(text: str) -> set[str]:
    """Split on non-alphanumeric chars, lower-case, drop empties."""
    return {t.lower() for t in re.split(r"[^a-zA-Z0-9]+", text) if t}


def _dept_matches(config_name: str, df_value: str) -> bool:
    """Return True if df_value is a reasonable match for config_name.

    Tier 1: exact match (lower-stripped).
    Tier 1b: underscore/hyphen → space normalisation.
    Tier 2: bidirectional token overlap — the shorter side's tokens must all
            appear in the longer side AND any extra tokens in the longer side
            must all be common noise words (dept, department, division, etc.).

            Handles:
              "Office_Engineering" ↔ "Office Engineering"   → match
              "Supervision"        ↔ "Supervision Dept"     → match (extra = "dept" = noise)
            Rejects:
              "Contract" ↔ "Procurement & Contract Administration"
              (extra = {"procurement", "administration"} — not noise)
    """
    a = config_name.strip().lower()
    b = df_value.strip().lower()

    if not a or not b:
        return False

    # Tier 1 — exact
    if a == b:
        return True

    # Tier 1b — underscore/hyphen → space normalisation
    a_norm = re.sub(r"[_\-]", " ", a).strip()
    b_norm = re.sub(r"[_\-]", " ", b).strip()
    if a_norm == b_norm:
        return True

    # Tier 2 — bidirectional token overlap
    ta = _tokenize(config_name)
    tb = _tokenize(df_value)
    if not ta or not tb:
        return False

    _NOISE = {"dept", "department", "division", "team", "unit", "office",
              "section", "group", "branch"}

    shorter, longer = (ta, tb) if len(ta) <= len(tb) else (tb, ta)

    # All tokens of the shorter side must appear in the longer side
    if not shorter.issubset(longer):
        return False

    # Extra tokens in the longer side must all be noise words
    extra = longer - shorter
    non_noise_extra = extra - _NOISE
    return len(non_noise_extra) == 0


def _source_files_for_dept(dept: dict, input_dir_hint: Path | None) -> set[str]:
    """Return a set of .xlsx filenames that live inside this dept's folder.

    Used as a last-resort fallback when no in-memory source-file mapping is
    available.  Prefer passing ``known_source_files`` to ``_build_dept_workbook``
    instead — that set is built from the already-loaded DataFrame and does NOT
    require the input folder to still exist on disk (i.e. it is safe to call
    after archiving has moved the files away).
    """
    folder_value = dept.get("folder", "").strip()
    if not folder_value:
        return set()

    dept_folder = Path(folder_value) if Path(folder_value).is_absolute() else (
        (input_dir_hint / folder_value) if input_dir_hint else Path(folder_value)
    )
    if not dept_folder.is_dir():
        return set()

    return {
        p.name for p in dept_folder.rglob("*.xlsx")
        if not p.name.startswith("~$")
        and "archive" not in [part.lower() for part in p.relative_to(dept_folder).parts[:-1]]
    }


def _source_files_for_dept_from_data(
    dept: dict,
    matched_df: pd.DataFrame,
    unmatched_sections: list[tuple[str, pd.DataFrame]],
    input_dir_hint: Path | None,
) -> set[str]:
    """Build the set of source filenames that belong to this department's folder.

    Unlike ``_source_files_for_dept`` this function derives the answer from the
    DataFrames that were already loaded into memory by the pipeline, so it is
    completely independent of whether the input files still exist on disk.

    The logic mirrors what ``discover_files`` + ``_collect_xlsx`` would have
    found: any file whose path resolves to inside this department's configured
    folder is considered to belong here.

    Falls back to a live disk scan only if no Source File column is present
    in the data (e.g. legacy flat-table runs).
    """
    folder_value = dept.get("folder", "").strip()
    if not folder_value:
        return set()

    dept_folder = Path(folder_value) if Path(folder_value).is_absolute() else (
        (input_dir_hint / folder_value) if input_dir_hint else Path(folder_value)
    )

    # ── Collect all source filenames recorded in the in-memory data ──────
    in_memory_sources: set[str] = set()

    if not matched_df.empty and "Source File" in matched_df.columns:
        for val in matched_df["Source File"].dropna().unique():
            in_memory_sources.add(Path(str(val)).name)

    for label, _ in unmatched_sections:
        in_memory_sources.add(Path(label).name)

    if not in_memory_sources:
        # No Source File info at all — fall back to live disk scan.
        return _source_files_for_dept(dept, input_dir_hint)

    # ── Filter to files that came from this dept's folder ────────────────
    # Strategy: reconstruct the original file paths by checking whether the
    # file *would have* been found inside dept_folder.  Since the files may
    # now be archived, we can't use .exists(); instead we resolve the folder
    # path and check the canonical folder prefix against known folder paths
    # recorded in the pipeline run.
    #
    # Simpler approach that works in practice: ask the pipeline to record the
    # full source path in matched_df.  Until that refactor lands, we cross-
    # reference using the dept config ``folder`` and the set of files that
    # the pipeline *would* have scanned from that folder.
    #
    # We achieve this without disk access by comparing the dept folder path
    # against the original file paths stored as "Source File" values in the
    # DataFrame.  Those values are just filenames (not full paths) today, so
    # we use a second strategy: check whether the file was the expected_file
    # for any employee in this department, OR whether it matches the folder
    # scan result when the folder still exists.

    # First try: if the folder still exists, use it (pre-archive scenario)
    if dept_folder.is_dir():
        disk_files = {
            p.name for p in dept_folder.rglob("*.xlsx")
            if not p.name.startswith("~$")
            and "archive" not in [part.lower() for part in p.relative_to(dept_folder).parts[:-1]]
        }
        if disk_files:
            return disk_files

    # Second try: derive from the full-path Source File values if they were
    # stored that way (future-proof path).
    result: set[str] = set()
    dept_folder_resolved = str(dept_folder.resolve()).lower().rstrip("/\\")

    if not matched_df.empty and "Source File" in matched_df.columns:
        for val in matched_df["Source File"].dropna().unique():
            p = Path(str(val))
            try:
                resolved = str(p.resolve()).lower()
            except Exception:
                resolved = str(p).lower()
            if resolved.startswith(dept_folder_resolved):
                result.add(p.name)

    for label, _ in unmatched_sections:
        p = Path(label)
        try:
            resolved = str(p.resolve()).lower()
        except Exception:
            resolved = str(p).lower()
        if resolved.startswith(dept_folder_resolved):
            result.add(p.name)

    # Third try (last resort): match by employee expected_file names from config
    if not result:
        expected = {
            emp.get("expected_file", "").strip()
            for emp in dept.get("employees", [])
            if emp.get("expected_file", "").strip()
        }
        result = expected & in_memory_sources

    return result


# ---------------------------------------------------------------------------
# Per-department workbook builder
# ---------------------------------------------------------------------------
def _build_dept_workbook(
    dept_name: str,
    dept: dict,
    matched_df: pd.DataFrame,
    unmatched_sections: list[tuple[str, pd.DataFrame]],
    week_label: str,
    input_dir_hint: Path | None = None,
    known_source_files: set[str] | None = None,
) -> "openpyxl.Workbook":  # noqa: F821
    """Build a styled workbook scoped to a single department.

    Parameters
    ----------
    known_source_files:
        Pre-computed set of .xlsx filenames that belong to this department's
        input folder.  Derived from the in-memory DataFrames by
        ``_source_files_for_dept_from_data`` — safe to use even after the
        input files have been archived.  When None the function falls back to
        a live disk scan (legacy behaviour, breaks after archiving).
    """
    from openpyxl import Workbook
    import main as _main  # noqa: PLC0415

    # ── Resolve the set of source filenames for this department ──────────
    # Prefer the pre-built in-memory set; fall back to disk scan only when
    # the caller has not provided one (e.g. direct calls from tests).
    if known_source_files is not None:
        dept_source_files = known_source_files
    else:
        dept_source_files = _source_files_for_dept_from_data(
            dept, matched_df, unmatched_sections, input_dir_hint
        )

    log.debug(
        "Dept '%s' source files (origin-matched): %s",
        dept_name,
        ", ".join(sorted(dept_source_files)) if dept_source_files else "(none)",
    )

    # ── Filter matched rows using three-tier matching ────────────────────
    if not matched_df.empty and "Department" in matched_df.columns:
        # Tier 1 & 2: Department column match
        dept_col_mask = matched_df["Department"].apply(
            lambda v: _dept_matches(dept_name, str(v)) if pd.notna(v) else False
        )

        # Tier 3: source file came from this dept's folder
        if dept_source_files and "Source File" in matched_df.columns:
            folder_mask = matched_df["Source File"].apply(
                lambda f: Path(str(f)).name in dept_source_files if pd.notna(f) else False
            )
            final_mask = dept_col_mask | folder_mask
        else:
            final_mask = dept_col_mask

        dept_df = matched_df[final_mask].copy()

        matched_count  = dept_col_mask.sum()
        folder_count   = (final_mask & ~dept_col_mask).sum() if dept_source_files else 0
        log.info(
            "Dept '%s': %d rows via Department column match, "
            "%d additional rows via folder-origin match (%d total)",
            dept_name, matched_count, folder_count, len(dept_df),
        )
    else:
        dept_df = pd.DataFrame()
        log.warning("Dept '%s': matched_df empty or missing Department column", dept_name)

    # ── Filter unmatched sections ────────────────────────────────────────
    dept_sections: list[tuple[str, pd.DataFrame]] = []
    for label, sec_df in unmatched_sections:
        label_fname = Path(label).name  # label is usually the source filename
        # Include if the label/filename matches this dept by name OR folder origin
        name_match = _dept_matches(dept_name, label)
        folder_match = label_fname in dept_source_files
        if name_match or folder_match:
            dept_sections.append((label, sec_df))
            log.debug(
                "Dept '%s': including unmatched section '%s' "
                "(name_match=%s, folder_match=%s)",
                dept_name, label, name_match, folder_match,
            )

    # ── Build workbook ───────────────────────────────────────────────────
    wb = Workbook()
    _main.build_details_sheet(wb, dept_df, dept_sections)
    _build_dept_summary_sheet(wb, dept_name, dept_df, week_label)

    # Summary first, Details second
    try:
        wb.move_sheet("Dept Summary", offset=-(len(wb.sheetnames) - 1))
    except Exception:
        pass
    wb.active = 0

    return wb


# ---------------------------------------------------------------------------
# Dept Summary sheet
# ---------------------------------------------------------------------------
def _build_dept_summary_sheet(
    wb,
    dept_name: str,
    dept_df: pd.DataFrame,
    week_label: str,
) -> None:
    """Lightweight summary sheet scoped to one department."""
    import main as _main  # noqa: PLC0415
    from openpyxl.styles import Font

    ws = wb.create_sheet("Dept Summary")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    Theme = _main.Theme

    ws["A1"] = f"{dept_name.replace('_', ' ')} — Weekly Department Report"
    ws["A1"].font = Theme.TITLE_FONT
    ws.merge_cells("A1:H1")

    ws["A2"] = f"Department Head View  ·  {week_label}  ·  CONFIDENTIAL"
    ws["A2"].font = Font(name=Theme.FONT, size=10, italic=True, color=Theme.DANGER)
    ws.merge_cells("A2:H2")

    if dept_df.empty:
        ws["A4"] = "No data was submitted for this department this week."
        ws["A4"].font = Theme.BODY_FONT
        _main.autofit_columns(ws)
        return

    rollup_df = (
        dept_df[dept_df["_Rollup Eligible"]]
        if "_Rollup Eligible" in dept_df.columns
        else dept_df
    )

    # ── KPI row ──────────────────────────────────────────────────────────
    total_hours = (
        pd.to_numeric(rollup_df["Hours Worked"], errors="coerce").sum()
        if "Hours Worked" in rollup_df.columns
        else 0
    )
    staff_count = (
        rollup_df["Staff Name"].nunique()
        if "Staff Name" in rollup_df.columns
        else None
    )

    kpis = [
        ("Total Rows",         len(dept_df)),
        ("Total Hours Logged", f"{total_hours:,.1f}"),
    ]
    if staff_count is not None:
        kpis.append(("Staff Reporting", staff_count))
        if staff_count > 0:
            kpis.append(("Avg Hours / Person", f"{total_hours / staff_count:,.1f}"))

    col = 1
    for label, value in kpis:
        ws.cell(row=4, column=col, value=label).font = Theme.KPI_LABEL_FONT
        ws.cell(row=5, column=col, value=value).font = Theme.KPI_VALUE_FONT
        col += 2

    row_cursor = 8

    # ── Staff breakdown ───────────────────────────────────────────────────
    if "Staff Name" in rollup_df.columns and not rollup_df.empty:
        agg_kwargs: dict = {"Rows": ("Staff Name", "count")}
        if "Hours Worked" in rollup_df.columns:
            agg_kwargs["Total Hours"] = ("Hours Worked", "sum")
        if "Project" in rollup_df.columns:
            agg_kwargs["Projects"] = ("Project", "nunique")
        if "Status" in rollup_df.columns:
            agg_kwargs["Completed Tasks"] = (
                "Status",
                lambda s: int(
                    (s.astype(str).str.lower().str.contains("complet")).sum()
                ),
            )

        staff_table = (
            rollup_df.groupby("Staff Name").agg(**agg_kwargs).reset_index()
        )
        if "Total Hours" in staff_table.columns:
            staff_table = staff_table.sort_values("Total Hours", ascending=False)

        ws.cell(row=row_cursor, column=1, value="Staff Breakdown").font = Theme.SUBTITLE_FONT
        row_cursor += 1
        row_cursor = _main.write_table_block(ws, row_cursor, staff_table)
        row_cursor += 2

    # ── Status breakdown ──────────────────────────────────────────────────
    if "Status" in rollup_df.columns and not rollup_df.empty:
        status_table = (
            rollup_df["Status"].astype(str).str.strip()
            .value_counts()
            .rename_axis("Status")
            .reset_index(name="Count")
        )
        status_table["% of Total"] = (
            (status_table["Count"] / status_table["Count"].sum() * 100).round(0)
        )
        ws.cell(row=row_cursor, column=1, value="Status Breakdown").font = Theme.SUBTITLE_FONT
        row_cursor += 1
        row_cursor = _main.write_table_block(ws, row_cursor, status_table)
        row_cursor += 2

    # ── Project breakdown ─────────────────────────────────────────────────
    if (
        "Project" in rollup_df.columns
        and "Hours Worked" in rollup_df.columns
        and rollup_df["Project"].notna().any()
    ):
        proj_table = (
            rollup_df.groupby("Project")["Hours Worked"]
            .sum()
            .sort_values(ascending=False)
            .round(1)
            .rename_axis("Project")
            .reset_index(name="Total Hours")
        )
        ws.cell(row=row_cursor, column=1, value="Hours by Project").font = Theme.SUBTITLE_FONT
        row_cursor += 1
        _main.write_table_block(ws, row_cursor, proj_table)

    _main.autofit_columns(ws)


# ---------------------------------------------------------------------------
# Main public entry point
# ---------------------------------------------------------------------------
def generate_dept_masters(
    matched_df: pd.DataFrame,
    unmatched_sections: list[tuple[str, pd.DataFrame]],
    departments: list[dict],
    output_dir: Path,
    week_label: str = "",
    input_dir: Path | None = None,
) -> list[DeptMasterResult]:
    """Generate one password-protected master Excel per department.

    Parameters
    ----------
    matched_df:
        The merged rollup-eligible DataFrame from the main pipeline.
    unmatched_sections:
        Raw sections from the same pipeline run.
    departments:
        The ``departments`` list from config.json.
    output_dir:
        Base output directory.  Per-dept files land in <output_dir>/dept_masters/.
    week_label:
        Human-readable week string embedded in the workbook title.
    input_dir:
        The pipeline's input_dir — used to resolve relative dept folder paths
        for the folder-origin matching fallback.

    Returns
    -------
    List of DeptMasterResult — one per department that has a dept_head_email.
    """
    if not week_label:
        from email_sender import week_label_for
        week_label = week_label_for()

    masters_dir = output_dir / "dept_masters"
    masters_dir.mkdir(parents=True, exist_ok=True)

    # ── Log what Department values are actually in the data ───────────────
    if not matched_df.empty and "Department" in matched_df.columns:
        unique_depts = (
            matched_df["Department"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )
        log.info(
            "Department column values in merged data: %s",
            ", ".join(f"'{d}'" for d in unique_depts) if unique_depts else "(none)",
        )
    else:
        log.warning(
            "Merged DataFrame is empty or has no Department column — "
            "dept masters will contain no matched rows"
        )

    results: list[DeptMasterResult] = []

    # ── Pre-compute per-department source-file sets from in-memory data ──
    # This is done BEFORE any disk access so it works correctly even when the
    # pipeline has already archived the input files (archive_inputs runs after
    # this function returns, but the ordering must not be relied on — having
    # the set derived from DataFrames makes it order-independent).
    dept_source_map: dict[str, set[str]] = {}
    for dept in departments:
        dept_name_key = dept.get("name", "").strip()
        if not dept_name_key:
            continue
        dept_source_map[dept_name_key] = _source_files_for_dept_from_data(
            dept, matched_df, unmatched_sections, input_dir
        )
        if dept_source_map[dept_name_key]:
            log.info(
                "Dept '%s' source files (from data): %s",
                dept_name_key,
                ", ".join(sorted(dept_source_map[dept_name_key])),
            )
        else:
            log.debug("Dept '%s': no source files identified in loaded data", dept_name_key)

    for dept in departments:
        dept_name       = dept.get("name", "").strip()
        dept_head_email = dept.get("dept_head_email", "").strip()
        dept_head_pass  = dept.get("dept_head_password", "").strip()

        if not dept_name:
            continue
        if not dept_head_email:
            log.debug("Skipping dept master for '%s': no dept_head_email.", dept_name)
            continue

        log.info("Generating dept master for: %s", dept_name)

        # ── Skip if this dept has no data in the current pipeline run ────
        # This prevents sending a blank "No data submitted" report to the
        # dept head when only some departments' files were present at the
        # time the pipeline ran (e.g. other depts haven't submitted yet).
        known_source_files = dept_source_map.get(dept_name, set())

        has_matched_rows = False
        if not matched_df.empty and "Department" in matched_df.columns:
            dept_col_mask = matched_df["Department"].apply(
                lambda v, dn=dept_name: _dept_matches(dn, str(v)) if pd.notna(v) else False
            )
            folder_mask = (
                matched_df["Source File"].apply(
                    lambda f, sf=known_source_files: Path(str(f)).name in sf
                    if pd.notna(f) else False
                )
                if known_source_files and "Source File" in matched_df.columns
                else pd.Series(False, index=matched_df.index)
            )
            has_matched_rows = (dept_col_mask | folder_mask).any()

        # Also check unmatched sections for this dept
        has_unmatched = any(
            _dept_matches(dept_name, Path(label).stem) or Path(label).name in known_source_files
            for label, _ in unmatched_sections
        )

        if not has_matched_rows and not has_unmatched:
            log.info(
                "Dept master for '%s' skipped — no files from this department "
                "were present in this pipeline run (files may not have been "
                "submitted yet, or were already archived from a previous run).",
                dept_name,
            )
            continue

        safe = _safe_name(dept_name)
        # Stable filename: dept + safe week label (no slashes/colons that break Windows paths)
        wl_safe = re.sub(r"[^a-zA-Z0-9_\-]", "_", week_label)
        final_path = masters_dir / f"{safe}_master_{wl_safe}.xlsx"

        try:
            wb = _build_dept_workbook(
                dept_name=dept_name,
                dept=dept,
                matched_df=matched_df,
                unmatched_sections=unmatched_sections,
                week_label=week_label,
                input_dir_hint=input_dir,
                known_source_files=dept_source_map.get(dept_name),
            )

            # Write to temp then atomically move (avoids partial-write corruption)
            fd, tmp_str = tempfile.mkstemp(
                suffix=".xlsx", dir=masters_dir, prefix=f".~{safe}_"
            )
            os.close(fd)
            tmp_path = Path(tmp_str)
            wb.save(tmp_path)

            encrypted = False
            if dept_head_pass:
                encrypted = password_protect(tmp_path, final_path, dept_head_pass)
                if encrypted:
                    tmp_path.unlink(missing_ok=True)
                else:
                    # Encryption failed — keep plain file
                    shutil.move(str(tmp_path), str(final_path))
            else:
                shutil.move(str(tmp_path), str(final_path))

            log.info(
                "Dept master saved: %s  (encrypted=%s, size=%s bytes)",
                final_path.name, encrypted,
                f"{final_path.stat().st_size:,}" if final_path.exists() else "?",
            )
            results.append(DeptMasterResult(
                dept_name=dept_name,
                output_path=final_path,
                encrypted=encrypted,
                dept_head_email=dept_head_email,
                error="",
            ))

        except Exception as exc:
            log.error("Failed to generate dept master for '%s': %s", dept_name, exc)
            for p in masters_dir.glob(f".~{safe}_*.xlsx"):
                try:
                    p.unlink()
                except Exception:
                    pass
            results.append(DeptMasterResult(
                dept_name=dept_name,
                output_path=Path(),
                encrypted=False,
                dept_head_email=dept_head_email,
                error=str(exc),
            ))

    return results
